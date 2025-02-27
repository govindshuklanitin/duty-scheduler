from flask import Flask, render_template, request, send_file, jsonify
from datetime import datetime, timedelta
import calendar
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

class DutyScheduler:
    def __init__(self):
        self.shifts = {
            'A': '06:00-14:00',
            'B': '14:00-22:00',
            'C': '22:00-06:00',
            'G': 'General',
            'R': 'Rest'
        }
        self.shift_rotation = {'A': 'C', 'C': 'B', 'B': 'A'}
        
    def get_day_name(self, year, month, day):
        return calendar.day_abbr[calendar.weekday(year, month, day)].upper()

    def get_next_shift(self, current_shift):
        return self.shift_rotation.get(current_shift, current_shift)

    def generate_schedule(self, employees_data, year, month):
        num_days = calendar.monthrange(year, month)[1]
        schedule = {}
        
        for emp in employees_data:
            schedule[emp['name']] = {
                'code': emp['code'],
                'post': emp['post'],
                'shifts': []
            }
            
            current_shift = emp['start_shift']
            rest_day = emp['rest_day']
            was_rest_day = False  # Flag to track if previous day was rest day
            
            for day in range(1, num_days + 1):
                # Check if it's a rest day
                if (day - 1) % 7 == rest_day:
                    schedule[emp['name']]['shifts'].append('R')
                    was_rest_day = True
                else:
                    # If previous day was rest day, change the shift according to rotation
                    if was_rest_day and current_shift != 'G':
                        current_shift = self.get_next_shift(current_shift)
                        was_rest_day = False
                    schedule[emp['name']]['shifts'].append(current_shift)
                    
        return schedule

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    data = request.get_json()
    year = int(data.get('year', datetime.now().year))
    month = int(data.get('month', datetime.now().month))
    employees_data = data.get('employees', [])
    
    scheduler = DutyScheduler()
    schedule = scheduler.generate_schedule(employees_data, year, month)
    
    return jsonify({
        'schedule': schedule,
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month]
    })

@app.route('/export', methods=['POST'])
def export():
    data = request.get_json()
    schedule = data.get('schedule', {})
    month = data.get('month')
    year = data.get('year')
    
    wb = Workbook()
    ws = wb.active
    
    # Styles
    header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")  # Dark grey color
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    thick_border = Border(
        left=Side(style='medium', color="000000"),
        right=Side(style='medium', color="000000"),
        top=Side(style='medium', color="000000"),
        bottom=Side(style='medium', color="000000")
    )
    
    # Calculate last column letter
    num_days = calendar.monthrange(year, month)[1]
    last_col = get_column_letter(num_days + 3)
    
    # Main Header - extend to last column
    ws.merge_cells(f'A1:C1')
    ws['A1'] = 'BAGASSE YARD'
    ws.merge_cells(f'D1:H1')
    ws['D1'] = 'SHIFT SCHEDULE'
    ws.merge_cells(f'I1:{last_col}1')
    ws[f'I1'] = f"{calendar.month_name[month].upper()} {year}"
    
    # Column Headers
    ws['A2'] = 'S.R'
    ws['B2'] = 'SUPERVISOR'
    ws['C2'] = 'CODE NO.'
    
    # Days and dates with header styling
    for day in range(1, num_days + 1):
        col = get_column_letter(day + 3)
        ws[f'{col}2'] = day
        ws[f'{col}3'] = DutyScheduler().get_day_name(year, month, day)
    
    # Apply header styles to ALL header cells (including dates and days)
    for row in range(1, 4):  # First 3 rows are headers
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply borders to ALL cells in the schedule
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if row < 4:  # Header rows
                cell.border = thick_border
            else:  # Data rows
                cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths for A4 paper - maximize space usage
    ws.column_dimensions['A'].width = 4  # S.R
    ws.column_dimensions['B'].width = 16  # Name
    ws.column_dimensions['C'].width = 8  # Code
    
    # Calculate optimal width for date columns to fill the page
    num_days = calendar.monthrange(year, month)[1]
    
    # Force wider width for date columns to fill the page
    date_column_width = 4.5  # Increased width for better visibility
    
    for col in range(4, num_days + 4):
        ws.column_dimensions[get_column_letter(col)].width = date_column_width
    
    # Set print settings for A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # Minimize margins to use maximum space
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    
    # Optimize row heights
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 16  # Standard height for all rows
    
    # Make header rows slightly taller
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    
    # Write schedule
    row = 4
    sr_no = 1
    
    # First write supervisors
    for emp_name, data in schedule.items():
        if data['post'].upper() == 'SUPERVISOR':
            ws[f'A{row}'] = sr_no
            ws[f'B{row}'] = emp_name
            ws[f'C{row}'] = data['code']
            
            for col, shift in enumerate(data['shifts'], start=4):
                ws.cell(row=row, column=col, value=shift)
            
            row += 1
            sr_no += 1
    
    # Then write helpers
    for emp_name, data in schedule.items():
        if data['post'].upper() == 'HELPER':
            ws[f'A{row}'] = sr_no
            ws[f'B{row}'] = emp_name
            ws[f'C{row}'] = data['code']
            
            for col, shift in enumerate(data['shifts'], start=4):
                ws.cell(row=row, column=col, value=shift)
            
            row += 1
            sr_no += 1
    
    # Color rest days in light red
    rest_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if cell.value == 'R':
                cell.fill = rest_fill
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'duty_schedule_{month}_{year}.xlsx'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
